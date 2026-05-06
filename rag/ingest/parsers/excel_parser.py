from __future__ import annotations

from .base import BaseParser, ParseResult, TextChunk


class ExcelParser(BaseParser):

    def can_parse(self, file_path: str) -> bool:
        return self._ext(file_path) in {".xlsx", ".xls", ".xlsm"}

    def parse(self, file_path: str) -> ParseResult:
        try:
            import openpyxl
        except ImportError:
            return ParseResult(
                source_path=file_path, doc_type="excel",
                errors=["openpyxl이 설치되지 않았습니다."]
            )

        chunks: list[TextChunk] = []
        errors: list[str] = []

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                rows_data = []
                headers: list[str] = []

                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    cells = [str(c) if c is not None else "" for c in row]
                    if row_idx == 0:
                        headers = cells
                    else:
                        if any(c.strip() for c in cells):
                            rows_data.append(cells)

                if not headers and not rows_data:
                    continue

                # 마크다운 테이블 형식으로 변환
                md_lines = []
                if headers:
                    md_lines.append("| " + " | ".join(headers) + " |")
                    md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
                for row in rows_data:
                    md_lines.append("| " + " | ".join(row) + " |")

                content = f"[시트: {sheet_name}]\n" + "\n".join(md_lines)
                chunks.append(
                    TextChunk(
                        content=content,
                        chunk_index=sheet_idx,
                        metadata={"sheet": sheet_name},
                    )
                )
            wb.close()
        except Exception as e:
            errors.append(str(e))

        return ParseResult(
            source_path=file_path,
            doc_type="excel",
            chunks=chunks,
            metadata={"sheet_count": len(chunks)},
            errors=errors,
        )
